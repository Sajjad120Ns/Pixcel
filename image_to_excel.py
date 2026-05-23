from get_pixels import GetPixel
import openpyxl as xl
from openpyxl.styles import PatternFill

class ImageToExcel():
    def __init__(self, image_path="image.jpg"):
        self.image_path = image_path

    def create_excel(self, output_name= "output.xlsx"):
        self.output_name = output_name
        self.wb = xl.Workbook()
        self.ws = self.wb.active

    def load_image(self):
        self.image = GetPixel(self.image_path)
        self.w, self.h = self.image.get_size()

    def adjust_cell_size(self, cell_size = 1):
        for row in range(1, self.h + 1):
            self.ws.row_dimensions[row].height = cell_size * 5.2
        for col in range(1, self.w + 1):
            col_letter = xl.utils.get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = cell_size

    def fill_excel(self):
        pixels = self.image.get_pixel()
        print("Coloring cells...")

        pixel_index = 0
        for row in range(1, self.h + 1):
            for col in range(1, self.w + 1):
                cell_color = pixels[pixel_index]
                cell = self.ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color= cell_color,
                                        end_color= cell_color,
                                         fill_type= "solid" 
                                         )
                pixel_index += 1
        self.image.close_image()
        print("Image convert successfully!")

    def save_file(self):
        self.wb.save(self.output_name)
        print(f"File saved as: {self.output_name}")
                


if __name__ == "__main__":
    new_image = ImageToExcel('images/example.jpg')
    new_image.create_excel()
    new_image.load_image()
    new_image.adjust_cell_size(cell_size=0.5)
    new_image.fill_excel()
    new_image.save_file()